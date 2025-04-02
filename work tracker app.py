from pathlib import Path
import logging
import sys
import os
import subprocess
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QLabel, QPushButton, QLineEdit, QDateEdit, QDoubleSpinBox, QSpinBox,
    QFileDialog, QMessageBox, QFrame, QGroupBox, QFormLayout, QTabWidget,
    QListWidget, QTableWidget, QTableWidgetItem, QScrollArea, QSplitter,
    QSpacerItem, QSizePolicy
)
from PyQt5.QtCore import Qt

# ----------------- Logging Setup -----------------
def setup_logging():
    # Use a folder in the user's home directory
    log_dir = Path.home() / "WorkTrackerLogs"
    log_dir.mkdir(exist_ok=True)
    log_path = log_dir / "tracker.log"

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)

    file_handler = logging.FileHandler(log_path, mode="a")
    stream_handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.info(f"Logging initialized. Log file: {log_path}")
    return logger

logger = setup_logging()

# ----------------- Clear Cache Function -----------------
def clear_cache():
    """Delete the log file to clear the cache."""
    log_file = Path.home() / "WorkTrackerLogs" / "tracker.log"
    if log_file.exists():
        try:
            log_file.unlink()  # Delete the file
            logger.info("Cache cleared. Log file removed.")
            return True
        except Exception as e:
            logger.error(f"Failed to clear cache: {e}")
            return False
    else:
        logger.info("No cache to clear. Log file does not exist.")
        return True

# ----------------- Helper Functions -----------------
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.getcwd()

def open_file(file_path):
    try:
        if sys.platform.startswith('win'):
            os.startfile(file_path)
        elif sys.platform.startswith('darwin'):
            subprocess.call(['open', file_path])
        else:
            subprocess.call(['xdg-open', file_path])
    except Exception as e:
        logger.error(f"Failed to open {file_path}: {e}")

def open_folder(file_path):
    folder = os.path.dirname(file_path)
    open_file(folder)

# ----------------- Excel Workbook Generation -----------------
def generate_workbook(data, start_date, end_date, cutoff_days, hourly_rate, employee_name):
    """
    Generates an Excel workbook with the following layout:

    - Row 1: "Date" label.
    - Row 2: Merged A2:H2 -> "Daily Recap ( - LA Office )"
    - Row 3: A3 = "Date", B3 = date range, D3 = "hrs", E3 = total hours (calculated as sum(Hr + (Min/60))).
    - Row 4: A4 = "Name", B4 = employee name, D4 = "8.00"
    - Row 5: A5 = "Department", B5 = department.
    - Row 6: Blank.
    - Row 7: Headers: Number, Daily Work Description, Hr, Min, Complete, Follow up, Supervisor Comments.
    - Row 8 onward: Data rows.

    A final "Total" sheet summarizes overall hours and total cost.
    """
    desired_cols = ["Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"]
    df = data.copy()
    existing_cols = [c for c in desired_cols if c in df.columns]
    df = df[existing_cols]

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    title_font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="005A9E", end_color="005A9E", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    highlight_font = Font(name="Calibri", bold=True, size=12, color="000000")
    label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    def format_daily_columns(ws):
        ws.column_dimensions["A"].width = 12   # "Date" label
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 60   # Daily Work Description (wider)
        ws.column_dimensions["D"].width = 8    # Hr (narrow)
        ws.column_dimensions["E"].width = 8    # Min (narrow)
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 60   # Supervisor Comments (wider)

    def add_daily_header(ws, seg, office="LA Office", employee=employee_name, department="Sales"):
        ws["A1"].value = "Date"
        ws.row_dimensions[1].height = 20

        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Daily Recap ( - {office} )"
        ws["A2"].font = title_font
        ws["A2"].fill = title_fill
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 30

        ws["A3"].value = "Date"
        ws["B3"].value = f"{seg[0].strftime('%Y-%m-%d')} to {seg[1].strftime('%Y-%m-%d')}"
        ws["D3"].value = "hrs"
        ws.row_dimensions[3].height = 20

        ws["A4"].value = "Name"
        ws["B4"].value = employee
        ws["D4"].value = "8.00"
        ws.row_dimensions[4].height = 20

        ws["A5"].value = "Department"
        ws["B5"].value = department
        ws.row_dimensions[5].height = 20

        ws.row_dimensions[6].height = 8

        headers = ["Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"]
        for col_idx, header_val in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col_idx)
            cell.value = header_val
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[7].height = 26

        ws.freeze_panes = "A8"

    segments = []
    current = start_date
    while current <= end_date:
        seg_end = current + timedelta(days=cutoff_days - 1)
        if seg_end > end_date:
            seg_end = end_date
        segments.append((current, seg_end))
        current = seg_end + timedelta(days=1)

    grand_total_hours = 0.0
    for seg in segments:
        sheet_name = seg[0].strftime("%m-%d-%Y")
        ws = wb.create_sheet(title=sheet_name)
        format_daily_columns(ws)
        add_daily_header(ws, seg)
        start_row = 8
        for row_vals in dataframe_to_rows(df, index=False, header=False):
            ws.append(row_vals)
        for r in range(start_row, ws.max_row + 1):
            ws.row_dimensions[r].height = 20
            fill_color = "F9F9F9" if r % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for c in range(1, 9):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.fill = row_fill
                if c == 3 or c == 8:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        seg_hours = 0.0
        for r in range(start_row, ws.max_row + 1):
            try:
                hr_val = float(ws.cell(row=r, column=3).value or 0)
            except:
                hr_val = 0
            try:
                min_val = float(ws.cell(row=r, column=4).value or 0)
            except:
                min_val = 0
            seg_hours += hr_val + (min_val / 60.0)
        ws["E3"].value = seg_hours
        grand_total_hours += seg_hours

    ws_total = wb.create_sheet(title="Total")
    ws_total.column_dimensions["A"].width = 2
    ws_total.column_dimensions["B"].width = 30
    ws_total.column_dimensions["C"].width = 15
    ws_total.column_dimensions["D"].width = 15
    ws_total.column_dimensions["E"].width = 15

    ws_total.merge_cells("B1:E1")
    ws_total["B1"].value = "Summary of All Sheets"
    ws_total["B1"].font = title_font
    ws_total["B1"].fill = title_fill
    ws_total["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_total.row_dimensions[1].height = 30
    ws_total.freeze_panes = "B3"

    total_cost = grand_total_hours * hourly_rate
    summary_data = [
        ("Total Hours Rendered", grand_total_hours),
        ("Hourly Rate", hourly_rate),
        ("Total (Rate * Hours)", total_cost),
        ("Cutoff Days", cutoff_days),
        ("Date Range", f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    ]
    row_idx = 3
    for label, val in summary_data:
        cell_label = ws_total.cell(row=row_idx, column=2)
        cell_label.value = label
        cell_label.font = Font(name="Calibri", bold=True, size=12)
        cell_label.alignment = Alignment(horizontal="right")
        cell_label.fill = label_fill
        cell_label.border = thin_border

        cell_val = ws_total.cell(row=row_idx, column=3)
        cell_val.value = val
        cell_val.border = thin_border
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        ws_total.row_dimensions[row_idx].height = 24
        row_idx += 1

    ws_total["C5"].fill = highlight_fill
    ws_total["C5"].font = highlight_font

    return wb

def export_to_pdf(workbook, pdf_path):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER

    doc = SimpleDocTemplate(pdf_path, pagesize=LETTER, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()

    custom_title = ParagraphStyle('customTitle', parent=styles['Title'],
                                  fontName='Helvetica-Bold', fontSize=20,
                                  textColor=colors.HexColor("#005A9E"))
    title = Paragraph("Tracker - Summary Report", custom_title)
    gen_date = Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"])
    elements.extend([title, Spacer(1, 12), gen_date, Spacer(1, 24)])

    total_sheet = workbook["Total"]
    overall_data = [
        ["Metric", "Value"],
        ["Total Hours Rendered", str(total_sheet["C3"].value)],
        ["Hourly Rate", str(total_sheet["C4"].value)],
        ["Total (Rate * Hours)", str(total_sheet["C5"].value)],
        ["Cutoff Days", str(total_sheet["C6"].value)],
        ["Date Range", str(total_sheet["C7"].value)]
    ]
    overall_table = Table(overall_data, colWidths=[220, 220])
    overall_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#007ACC")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 12),
        ("BOTTOMPADDING", (0,0), (-1,0), 10),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
    ]))
    elements.extend([Paragraph("Overall Summary", styles["Heading2"]), Spacer(1, 6), overall_table, Spacer(1, 24)])

    sheet_data = [["Sheet Name", "Hours Rendered"]]
    for sheet in workbook.worksheets:
        if sheet.title == "Total":
            continue
        try:
            hours = float(sheet["E3"].value or 0)
        except:
            hours = 0
        sheet_data.append([sheet.title, str(hours)])
    sheet_table = Table(sheet_data, colWidths=[250, 100])
    sheet_table_style = TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#007ACC")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 12),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
    ])
    for i in range(1, len(sheet_data)):
        bg_color = colors.lightgrey if i % 2 == 0 else colors.whitesmoke
        sheet_table_style.add("BACKGROUND", (0, i), (-1, i), bg_color)
    sheet_table.setStyle(sheet_table_style)
    elements.extend([Paragraph("Sheet Breakdown", styles["Heading2"]), Spacer(1, 6), sheet_table])

    doc.build(elements)
    return pdf_path

# ----------------- PyQt5 GUI Application -----------------
class TrackerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.logger = logging.getLogger(__name__)
        self.setWindowTitle("Work Tracker")
        self.resize(1200, 800)
        self.df = None
        self.workbook = None
        self.export_history = []
        self.setup_ui()

    def setup_ui(self):
        splitter = QSplitter(Qt.Horizontal)
        self.setCentralWidget(splitter)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 7)

        # Left Panel
        self.left_frame = QFrame()
        self.left_frame.setStyleSheet("QFrame { background-color: #F7F7F7; border: 1px solid #999; border-radius: 8px; }")
        left_layout = QVBoxLayout(self.left_frame)
        left_layout.setContentsMargins(12, 12, 12, 12)
        left_layout.setSpacing(14)

        title_label = QLabel("Work Track")
        title_label.setStyleSheet("font-size: 22px; font-weight: bold; color: #2A2A60;")
        left_layout.addWidget(title_label, 0, Qt.AlignHCenter)

        # Import Work Log Group
        file_group = QGroupBox("Import Work Log")
        file_group.setStyleSheet("""
            QGroupBox { font-size: 12pt; color: #2A2A60; border: 1px solid #cccccc; border-radius: 6px; margin-top: 6px; }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top center; padding: 0 8px; }
        """)
        file_layout = QHBoxLayout(file_group)
        self.file_line_edit = QLineEdit()
        self.file_line_edit.setReadOnly(True)
        self.file_line_edit.setStyleSheet("background-color: #FFFFFF;")
        browse_btn = QPushButton("Browse")
        browse_btn.setStyleSheet("QPushButton { background-color: #333333; color: white; padding: 6px; } QPushButton:hover { background-color: #555555; }")
        browse_btn.clicked.connect(self.browse_file)
        preview_btn = QPushButton("Preview")
        preview_btn.setStyleSheet("QPushButton { background-color: #333333; color: white; padding: 6px; } QPushButton:hover { background-color: #555555; }")
        preview_btn.clicked.connect(self.preview_file)
        file_layout.addWidget(self.file_line_edit)
        file_layout.addWidget(browse_btn)
        file_layout.addWidget(preview_btn)
        left_layout.addWidget(file_group)

        # Import History Group
        history_group = QGroupBox("Import History")
        history_group.setStyleSheet("""
            QGroupBox { font-size: 10pt; color: #2A2A60; border: 1px solid #cccccc; border-radius: 6px; margin-top: 6px; }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top center; padding: 0 8px; }
        """)
        history_layout = QVBoxLayout(history_group)
        self.history_list = QListWidget()
        history_layout.addWidget(self.history_list)
        del_import_btn = QPushButton("Delete Previous Imports")
        del_import_btn.setStyleSheet("QPushButton { background-color: #333333; color: white; padding: 4px; } QPushButton:hover { background-color: #555555; }")
        del_import_btn.clicked.connect(self.delete_import_history)
        history_layout.addWidget(del_import_btn)
        left_layout.addWidget(history_group)

        refresh_btn = QPushButton("Refresh All")
        refresh_btn.setStyleSheet("QPushButton { background-color: #333333; color: white; padding: 6px; } QPushButton:hover { background-color: #555555; }")
        refresh_btn.clicked.connect(self.refresh_all)
        left_layout.addWidget(refresh_btn)
        
        # Clear Cache Button
        clear_cache_btn = QPushButton("Clear Cache")
        clear_cache_btn.setStyleSheet("QPushButton { background-color: #AA0000; color: white; padding: 6px; } QPushButton:hover { background-color: #CC0000; }")
        clear_cache_btn.clicked.connect(self.clear_cache_ui)
        left_layout.addWidget(clear_cache_btn)

        # Instruction Group
        instr_container = QGroupBox("Demo Instruction")
        instr_container.setStyleSheet("""
            QGroupBox { background-color: #F0F0F0; font-size: 10pt; color: #000000; border: 1px solid #CCCCCC; border-radius: 6px; margin-top: 6px; padding: 8px; }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top center; padding: 0 10px; }
        """)
        instr_label = QLabel(
            "1. Import a work log file using 'Browse' (txt/csv/xlsx) and then click 'Preview'.<br/>"
            "2. Set generation parameters in the lower-right panel.<br/>"
            "3. Click the large 'Generate XLSX' button to create a multi-sheet Excel report; new generations are appended.<br/>"
            "4. Use the export buttons below to save your report as Excel (black), PDF (dark gray), or CSV (dark blue).<br/>"
            "5. Click 'Explore New Sheet' to add more logs as new sheets.<br/>"
            "6. The real-time filter above the preview tabs lets you jump to a sheet by date (MM-DD-YYYY).<br/>"
            "7. The Excel Total sheet uses column B for labels (wide for readability)."
        )
        instr_label.setWordWrap(True)
        instr_container_layout = QVBoxLayout(instr_container)
        instr_container_layout.addWidget(instr_label)
        left_layout.addWidget(instr_container)
        left_layout.addItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # Export History Group
        export_history_group = QGroupBox("Export History")
        export_history_group.setStyleSheet("""
            QGroupBox { font-size: 10pt; color: #2A2A60; border: 1px solid #cccccc; border-radius: 6px; margin-top: 6px; }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top center; padding: 0 8px; }
        """)
        export_history_layout = QVBoxLayout(export_history_group)
        self.export_history_list = QListWidget()
        export_history_layout.addWidget(self.export_history_list)
        left_layout.addWidget(export_history_group)

        left_layout.addStretch()
        splitter.addWidget(self.left_frame)

        # Right Panel
        self.right_frame = QFrame()
        self.right_frame.setStyleSheet("QFrame { background-color: #FFFDFB; border: 1px solid #D8D2CA; border-radius: 8px; }")
        right_layout = QVBoxLayout(self.right_frame)
        right_layout.setContentsMargins(10, 10, 10, 10)
        right_layout.setSpacing(12)

        # Top: Title + Filter
        top_right = QWidget()
        top_right_layout = QHBoxLayout(top_right)
        top_right_layout.setContentsMargins(0, 0, 0, 0)
        preview_title = QLabel("Spreadsheet Preview")
        preview_title.setStyleSheet("font-size: 20px; font-weight: bold; color: #2A2A60;")
        preview_title.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        top_right_layout.addWidget(preview_title)
        self.tab_filter_edit = QLineEdit()
        self.tab_filter_edit.setPlaceholderText("Enter tab date (MM-DD-YYYY) to jump")
        self.tab_filter_edit.textChanged.connect(self.filter_tabs)
        top_right_layout.addWidget(self.tab_filter_edit)
        right_layout.addWidget(top_right)

        # Middle: Tab Widget
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.setStyleSheet("""
            QTabBar::tab { min-width: 100px; padding: 8px; margin: 2px; }
            QTabBar::tab:selected { background: #D8E4F0; border: 1px solid #999; border-radius: 6px; }
        """)
        self.raw_data_tab = QWidget()
        self.raw_data_layout = QVBoxLayout(self.raw_data_tab)
        self.raw_data_table = QTableWidget()
        self.raw_data_table.verticalHeader().setVisible(False)
        self.raw_data_table.setAlternatingRowColors(True)
        self.raw_data_table.setStyleSheet("""
            QTableWidget { background-color: #FFFFFF; }
            QTableWidget::item { padding: 4px; }
            QTableWidget::item:selected { background-color: #ADD8E6; }
        """)
        scroll_raw = QScrollArea()
        scroll_raw.setWidgetResizable(True)
        scroll_raw.setWidget(self.raw_data_table)
        self.raw_data_layout.addWidget(scroll_raw)
        self.tabs.addTab(self.raw_data_tab, "Raw Data")
        right_layout.addWidget(self.tabs)

        # Bottom: Generation Parameters + Export Buttons
        gen_params_group = QGroupBox("Generation Parameters")
        gen_params_group.setStyleSheet("""
            QGroupBox { background-color: #F0F0F0; font-size: 12pt; color: #2A2A60; border: 1px solid #666666; border-radius: 4px; padding: 8px; margin-top: 6px; }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top center; padding: 0 10px; }
        """)
        gen_params_layout = QFormLayout(gen_params_group)
        gen_params_layout.setHorizontalSpacing(20)
        gen_params_layout.setVerticalSpacing(12)

        self.start_date_edit2 = QDateEdit(calendarPopup=True)
        self.start_date_edit2.setDisplayFormat("yyyy-MM-dd")
        self.start_date_edit2.setDate(QtCore.QDate.currentDate())
        self.end_date_edit2 = QDateEdit(calendarPopup=True)
        self.end_date_edit2.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit2.setDate(QtCore.QDate.currentDate())
        self.cutoff_spin2 = QSpinBox()
        self.cutoff_spin2.setRange(1, 60)
        self.cutoff_spin2.setValue(7)
        self.rate_spin2 = QDoubleSpinBox()
        self.rate_spin2.setRange(0.0, 9999.99)
        self.rate_spin2.setDecimals(2)
        self.rate_spin2.setValue(8.00)
        gen_params_layout.addRow("Start Date:", self.start_date_edit2)
        gen_params_layout.addRow("End Date:", self.end_date_edit2)
        gen_params_layout.addRow("Days/Sheet:", self.cutoff_spin2)
        gen_params_layout.addRow("Hourly Rate:", self.rate_spin2)

        self.user_name_edit2 = QLineEdit()
        self.user_name_edit2.setPlaceholderText("Enter Employee Name")
        gen_params_layout.addRow("Employee Name:", self.user_name_edit2)

        generate_btn = QPushButton("Generate XLSX")
        generate_btn.setStyleSheet("""
            QPushButton { background-color: #333333; color: white; font-size: 14pt; padding: 10px; }
            QPushButton:hover { background-color: #555555; }
        """)
        generate_btn.clicked.connect(self.generate_spreadsheet)
        gen_params_layout.addRow(generate_btn)

        export_buttons_layout = QHBoxLayout()
        export_btn_style = """
            QPushButton { background-color: black; color: white; padding: 6px; text-transform: uppercase; }
            QPushButton:hover { background-color: #666666; }
        """
        export_excel_btn = QPushButton("Export Excel")
        export_excel_btn.setStyleSheet(export_btn_style)
        export_excel_btn.clicked.connect(self.export_as_excel)
        export_pdf_btn = QPushButton("Export PDF")
        export_pdf_btn.setStyleSheet(export_btn_style)
        export_pdf_btn.clicked.connect(self.export_as_pdf)
        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.setStyleSheet(export_btn_style)
        export_csv_btn.clicked.connect(self.export_as_csv)
        export_buttons_layout.setSpacing(10)
        export_buttons_layout.addWidget(export_excel_btn)
        export_buttons_layout.addWidget(export_pdf_btn)
        export_buttons_layout.addWidget(export_csv_btn)
        gen_params_layout.addRow(export_buttons_layout)

        explore_btn = QPushButton("Explore New Sheet")
        explore_btn.setStyleSheet("""
            QPushButton { background-color: darkgray; color: white; padding: 8px; font-weight: bold; }
            QPushButton:hover { background-color: #aaaaaa; }
        """)
        explore_btn.clicked.connect(self.explore_new_sheet)
        gen_params_layout.addRow(explore_btn)

        scroll_gen = QScrollArea()
        scroll_gen.setWidgetResizable(True)
        scroll_gen.setWidget(gen_params_group)
        right_layout.addWidget(scroll_gen)

        splitter.addWidget(self.right_frame)

    def delete_import_history(self):
        reply = QMessageBox.question(
            self, "Delete Imports", "Are you sure you want to delete all previous imports?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.history_list.clear()

    def clear_cache_ui(self):
        # Clear the log cache and show a message box.
        if clear_cache():
            QMessageBox.information(self, "Clear Cache", "Cache cleared successfully.")
        else:
            QMessageBox.critical(self, "Clear Cache", "Failed to clear cache.")

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Work Log File", "",
            "Text Files (*.txt);;CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)"
        )
        if file_path:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.file_line_edit.setText(file_path)
            self.df = None
            self.raw_data_table.clear()
            self.history_list.addItem(f"{file_path} ({timestamp})")
            self.logger.info(f"File loaded: {file_path}")

    def preview_file(self):
        file_path = self.file_line_edit.text()
        if not file_path or not os.path.isfile(file_path):
            self.logger.error("Invalid file path")
            QMessageBox.critical(self, "Error", "Please select a valid file.")
            return
        try:
            if file_path.lower().endswith(".txt"):
                df = pd.read_csv(file_path, sep="\t")
            elif file_path.lower().endswith(".csv"):
                df = pd.read_csv(file_path)
            elif file_path.lower().endswith(".xlsx"):
                df = pd.read_excel(file_path)
            else:
                df = pd.read_csv(file_path)
            self.df = df
            preview_df = df.head(30)
            self.populate_tablewidget(self.raw_data_table, preview_df)
            self.logger.info(f"File previewed: {file_path}")
        except Exception as e:
            self.logger.error(f"Failed to preview file: {e}")
            QMessageBox.critical(self, "Error", f"Failed to preview file:\n{e}")

    def populate_tablewidget(self, table: QTableWidget, df: pd.DataFrame):
        table.clear()
        table.setRowCount(len(df))
        table.setColumnCount(len(df.columns))
        table.setHorizontalHeaderLabels([str(c) for c in df.columns])
        for i in range(len(df)):
            for j in range(len(df.columns)):
                table.setItem(i, j, QTableWidgetItem(str(df.iat[i, j])))
        table.resizeColumnsToContents()
        table.resizeRowsToContents()

    def filter_tabs(self):
        text = self.tab_filter_edit.text().strip().lower()
        for i in range(self.tabs.count()):
            tab_name = self.tabs.tabText(i).lower()
            if text in tab_name:
                self.tabs.setCurrentIndex(i)
                return

    def generate_spreadsheet(self):
        if self.df is None:
            reply = QMessageBox.question(
                self, "No Data Imported", "No data has been imported. Generate an empty template?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
            else:
                self.df = pd.DataFrame()

        start_date = self.start_date_edit2.date().toPyDate()
        end_date = self.end_date_edit2.date().toPyDate()
        if end_date < start_date:
            self.logger.error("Invalid date range")
            QMessageBox.critical(self, "Error", "End date must not be before start date.")
            return

        cutoff_days = self.cutoff_spin2.value()
        hourly_rate = self.rate_spin2.value()
        employee_name = self.user_name_edit2.text().strip()
        if not employee_name:
            self.logger.error("Employee name not provided")
            QMessageBox.warning(self, "Input Required", "Please enter an employee name.")
            return

        try:
            new_wb = generate_workbook(self.df, start_date, end_date, cutoff_days, hourly_rate, employee_name)
            self.workbook = new_wb
            self.logger.info("Spreadsheet generated in memory.")
        except Exception as e:
            self.logger.error(f"Failed to generate spreadsheet: {e}")
            QMessageBox.critical(self, "Error", f"Failed to generate spreadsheet:\n{e}")
            return

        self.populate_preview_tabs()
        QMessageBox.information(self, "Success", "Spreadsheet generated in memory.\nYou can now export it.")

    def populate_preview_tabs(self):
        while self.tabs.count() > 1:
            self.tabs.removeTab(self.tabs.count() - 1)
        if not self.workbook:
            return
        for sheet in self.workbook.worksheets:
            if sheet.title == "Raw Data":
                continue
            tab = QWidget()
            layout = QVBoxLayout(tab)
            table = QTableWidget()
            table.verticalHeader().setVisible(False)
            table.setAlternatingRowColors(True)
            table.setStyleSheet("""
                QTableWidget { background-color: #FFFFFF; }
                QTableWidget::item { padding: 4px; }
                QTableWidget::item:selected { background-color: #ADD8E6; }
            """)
            layout.addWidget(table)
            self.tabs.addTab(tab, sheet.title)
            data_list = []
            headers = []
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c) if c is not None else "" for c in row]
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True):
                row_vals = [str(c) if c is not None else "" for c in row]
                data_list.append(row_vals)
            table.setRowCount(len(data_list))
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            for i in range(len(data_list)):
                for j in range(len(headers)):
                    table.setItem(i, j, QTableWidgetItem(data_list[i][j]))
            table.resizeColumnsToContents()
            table.resizeRowsToContents()

    def explore_new_sheet(self):
        new_file, _ = QFileDialog.getOpenFileName(
            self, "Select New Work Log File", "",
            "Text Files (*.txt);;CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)"
        )
        if not new_file:
            return
        try:
            if new_file.lower().endswith(".txt"):
                new_df = pd.read_csv(new_file, sep="\t")
            elif new_file.lower().endswith(".csv"):
                new_df = pd.read_csv(new_file)
            elif new_file.lower().endswith(".xlsx"):
                new_df = pd.read_excel(new_file)
            else:
                new_df = pd.read_csv(new_file)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load new file:\n{e}")
            return
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.history_list.addItem(f"{new_file} ({timestamp})")

        start_date = self.start_date_edit2.date().toPyDate()
        end_date = self.end_date_edit2.date().toPyDate()
        cutoff_days = self.cutoff_spin2.value()
        hourly_rate = self.rate_spin2.value()
        employee_name = self.user_name_edit2.text().strip()

        try:
            new_wb = generate_workbook(new_df, start_date, end_date, cutoff_days, hourly_rate, employee_name)
            if self.workbook is None:
                self.workbook = new_wb
            else:
                for sheet in new_wb.worksheets:
                    if sheet.title != "Total":
                        self.workbook._sheets.append(sheet)
                self.workbook.remove(self.workbook["Total"])
                self.workbook._sheets.append(new_wb["Total"])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add new sheet:\n{e}")
            return

        self.populate_preview_tabs()
        QMessageBox.information(self, "Success", "New sheet added from selected file.")

    def refresh_all(self):
        self.file_line_edit.clear()
        self.df = None
        self.raw_data_table.clear()
        self.history_list.clear()
        self.workbook = None
        while self.tabs.count() > 1:
            self.tabs.removeTab(self.tabs.count() - 1)
        self.export_history_list.clear()
        QMessageBox.information(self, "Refresh", "All inputs have been refreshed.")

    def export_as_excel(self):
        if not self.workbook:
            self.logger.error("No spreadsheet generated")
            QMessageBox.critical(self, "Error", "No spreadsheet generated.")
            return

        base_path = get_base_path()
        export_dir = os.path.join(base_path, "Work Tracker", "XLSX")
        os.makedirs(export_dir, exist_ok=True)
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", export_dir, "Excel Files (*.xlsx)")
        if file_path:
            try:
                self.workbook.save(file_path)
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.logger.info(f"Excel file saved: {file_path}")
                self.export_history.append(f"{file_path} ({timestamp})")
                self.export_history_list.addItem(f"Exported XLSX: {file_path} ({timestamp})")
                QMessageBox.information(self, "Success", f"Excel file saved at:\n{file_path}")
                open_file(file_path)
                open_folder(file_path)
            except Exception as e:
                self.logger.error(f"Failed to save Excel file: {e}")
                QMessageBox.critical(self, "Error", f"Failed to save Excel file:\n{e}")

    def export_as_pdf(self):
        if not self.workbook:
            self.logger.error("No spreadsheet generated")
            QMessageBox.critical(self, "Error", "No spreadsheet generated.")
            return

        base_path = get_base_path()
        export_dir = os.path.join(base_path, "Work Tracker", "PDF")
        os.makedirs(export_dir, exist_ok=True)
        file_path, _ = QFileDialog.getSaveFileName(self, "Save PDF File", export_dir, "PDF Files (*.pdf)")
        if file_path:
            try:
                export_to_pdf(self.workbook, file_path)
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.logger.info(f"PDF file saved: {file_path}")
                self.export_history.append(f"{file_path} ({timestamp})")
                self.export_history_list.addItem(f"Exported PDF: {file_path} ({timestamp})")
                QMessageBox.information(self, "Success", f"PDF file saved at:\n{file_path}")
                open_file(file_path)
                open_folder(file_path)
            except Exception as e:
                self.logger.error(f"Failed to export PDF: {e}")
                QMessageBox.critical(self, "Error", f"Failed to export PDF:\n{e}")

    def export_as_csv(self):
        if self.df is None:
            self.logger.error("No data loaded")
            QMessageBox.critical(self, "Error", "No data loaded.")
            return

        base_path = get_base_path()
        export_dir = os.path.join(base_path, "Work Tracker", "CSV")
        os.makedirs(export_dir, exist_ok=True)
        file_path, _ = QFileDialog.getSaveFileName(self, "Save CSV File", export_dir, "CSV Files (*.csv)")
        if file_path:
            try:
                self.df.to_csv(file_path, index=False)
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.logger.info(f"CSV file saved: {file_path}")
                self.export_history.append(f"{file_path} ({timestamp})")
                self.export_history_list.addItem(f"Exported CSV: {file_path} ({timestamp})")
                QMessageBox.information(self, "Success", f"CSV file saved at:\n{file_path}")
                open_file(file_path)
                open_folder(file_path)
            except Exception as e:
                self.logger.error(f"Failed to export CSV: {e}")
                QMessageBox.critical(self, "Error", f"Failed to export CSV:\n{e}")

    def search_in_preview(self):
        search_term = self.tab_filter_edit.text().lower()
        if not search_term:
            return
        current_tab = self.tabs.currentWidget()
        if current_tab is None:
            return
        table = current_tab.findChild(QTableWidget)
        if table:
            for i in range(table.rowCount()):
                for j in range(table.columnCount()):
                    item = table.item(i, j)
                    if item and search_term in item.text().lower():
                        item.setBackground(QtGui.QColor("yellow"))
                    else:
                        item.setBackground(QtGui.QColor("white"))

def main():
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    file_handler = logging.FileHandler('tracker.log')
    stream_handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = TrackerApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
